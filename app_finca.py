import streamlit as st
import pandas as pd
import mysql.connector
import io
import requests 
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. Configuración de la página
st.set_page_config(page_title="Sistema Agrícola V3.0", layout="wide")

# --- MÓDULO 1: SISTEMA DE SEGURIDAD Y LOGIN ---
if 'autorizado' not in st.session_state:
    st.session_state['autorizado'] = False

# Pantalla de Login
if not st.session_state['autorizado']:
    st.markdown("<h1 style='text-align: center; color: #1e3f20;'>🔒 Acceso Restringido</h1>", unsafe_allow_html=True)
    st.markdown("<p style='text-align: center;'>Sistema Central de Monitoreo - Finca 300 Hectáreas</p>", unsafe_allow_html=True)
    
    col_izq, col_centro, col_der = st.columns([1, 1, 1])
    with col_centro:
        with st.form("login_form"):
            usuario = st.text_input("Usuario")
            clave = st.text_input("Contraseña", type="password")
            btn_login = st.form_submit_button("Entrar al Sistema", use_container_width=True)
            
            if btn_login:
                if usuario == "admin" and clave == "finca2026":
                    st.session_state['autorizado'] = True
                    st.rerun() 
                else:
                    st.error("❌ Credenciales incorrectas. Acceso denegado.")
                    
    st.stop()

# ==============================================================================
# SI EL CÓDIGO LLEGA HASTA AQUÍ, SIGNIFICA QUE EL USUARIO ESTÁ AUTORIZADO
# ==============================================================================

st.title("🌱 Panel de Control Central - Finca V3.0")
st.markdown("---")
st.success("✅ Conexión segura establecida como Administrador.")

def obtener_conexion():
    return mysql.connector.connect(
        host="bpvhrmazb58ojyt1ynth-mysql.services.clever-cloud.com",
        user="uunwlo7t4x1ihcw9",
        password="3OEBQ5ERzoCLY6u9SvXG", 
        database="bpvhrmazb58ojyt1ynth"
    )

# --- MÓDULO 3: ALERTAS DIRECTAS AL CELULAR VÍA BLYNK ---
def disparar_alerta_movil(id_parcela, humedad, agua):
    # Token oficial y cuenta verificada
    TOKEN_BLYNK = "wD5j1l-ymUsNVwJOy4jDw5zS0-fiylZS" 
    EVENTO = "alerta_riego"
    
    descripcion = f"URGENTE: {id_parcela} bajo a {humedad}%. Requiere {agua}m3 de agua."
    
    url_blynk = f"https://blynk.cloud/external/api/logEvent?token={TOKEN_BLYNK}&event={EVENTO}&description={descripcion}"
    
    try:
        requests.get(url_blynk, timeout=3)
        st.toast(f"📱 Señal enviada al teléfono móvil. {id_parcela} requiere riego.", icon="🚨")
    except Exception as e:
        st.warning("No se pudo conectar con el servidor móvil de Blynk, pero el registro se guardó.")

# --- CEREBRO AUTOMÁTICO ---
def calcular_estado_y_agua(cultivo, humedad, hectareas):
    cultivo_limpio = str(cultivo).lower()
    if "arroz" in cultivo_limpio:
        target = 75
    elif "frijol" in cultivo_limpio:
        target = 40
    elif "maíz" in cultivo_limpio or "maiz" in cultivo_limpio:
        target = 50
    else:
        target = 45
        
    estado = "Óptimo" if humedad >= target else "Requiere Riego (Crítico)"
    deficit = target - humedad
    agua_m3 = (deficit * hectareas * 8) if deficit > 0 else 0
        
    return estado, agua_m3

# --- PANEL LATERAL (Registro Manual) ---
st.sidebar.header("➕ Registrar Nueva Parcela")
with st.sidebar.form("formulario_parcela", clear_on_submit=True):
    id_parcela = st.text_input("ID de la Parcela (Ej: PAR-006)").strip()
    sector = st.selectbox("Sector", ["Norte", "Sur", "Este", "Oeste", "Central"])
    hectareas = st.number_input("Hectáreas", min_value=1, max_value=300, value=10)
    cultivo = st.text_input("Tipo de Cultivo (Ej: Arroz, Caña)").strip()
    humedad = st.slider("Humedad del Suelo (%)", 0, 100, 50)
    
    if st.form_submit_button("Guardar en Nube"):
        if id_parcela and cultivo:
            estado_calc, agua_calc = calcular_estado_y_agua(cultivo, humedad, hectareas)
            conn = obtener_conexion()
            cursor = conn.cursor()
            cursor.execute("INSERT INTO registro_parcelas (id_parcela, sector, hectareas, cultivo, humedad_suelo_pct, estado) VALUES (%s, %s, %s, %s, %s, %s)", 
                           (id_parcela, sector, hectareas, cultivo, humedad, estado_calc))
            conn.commit()
            cursor.close()
            conn.close()
            
            if "Crítico" in estado_calc or "Requiere Riego" in estado_calc:
                disparar_alerta_movil(id_parcela, humedad, agua_calc)
            st.sidebar.success(f"¡{id_parcela} procesada con éxito!")
        else:
            st.sidebar.warning("⚠️ Rellena todos los campos.")

# --- LECTURA DE DATOS Y PESTAÑAS ---
try:
    conn = obtener_conexion()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT id_parcela, sector, hectareas, cultivo, humedad_suelo_pct, estado FROM registro_parcelas")
    resultados = cursor.fetchall()
    cursor.close()
    conn.close()
    
    if resultados:
        df = pd.DataFrame(resultados)
        df['Agua Necesaria (m³)'] = df.apply(lambda r: calcular_estado_y_agua(r['cultivo'], r['humedad_suelo_pct'], r['hectareas'])[1], axis=1)
        
        pestana1, pestana2 = st.tabs(["📊 Editor Interactivo (Live Excel)", "📥 Reportes Ejecutivos"])
        
        # --- MÓDULO 2: EDICIÓN EN VIVO ---
        with pestana1:
            st.subheader("📝 Cuadrícula Dinámica de Datos")
            st.info("💡 Haz doble clic en cualquier celda para editar (ej: cambia la humedad). Luego presiona el botón para sincronizar con el servidor.")
            
            df_editado = st.data_editor(df, use_container_width=True, hide_index=True)
            
            if st.button("🔄 Sincronizar Cambios con la Base de Datos", type="primary"):
                try:
                    conn = obtener_conexion()
                    cursor = conn.cursor()
                    
                    for index, fila in df_editado.iterrows():
                        estado_nuevo, _ = calcular_estado_y_agua(fila['cultivo'], fila['humedad_suelo_pct'], fila['hectareas'])
                        
                        sql = """INSERT INTO registro_parcelas (id_parcela, sector, hectareas, cultivo, humedad_suelo_pct, estado) 
                                 VALUES (%s, %s, %s, %s, %s, %s)
                                 ON DUPLICATE KEY UPDATE 
                                 sector=VALUES(sector), hectareas=VALUES(hectareas), cultivo=VALUES(cultivo), 
                                 humedad_suelo_pct=VALUES(humedad_suelo_pct), estado=VALUES(estado)"""
                        
                        cursor.execute(sql, (fila['id_parcela'], fila['sector'], fila['hectareas'], fila['cultivo'], fila['humedad_suelo_pct'], estado_nuevo))
                    
                    conn.commit()
                    cursor.close()
                    conn.close()
                    st.success("¡Base de datos sincronizada con éxito! Todos los cambios han sido guardados.")
                    st.rerun() 
                except Exception as e:
                    st.error(f"Error al sincronizar: {e}")

        # --- PESTAÑA 2: DESCARGAS ---
        with pestana2:
            st.subheader("🔍 Exportador de Datos")
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Reporte_Finca')
                worksheet = writer.sheets['Reporte_Finca']
                color_enc = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") 
                fuente_enc = Font(color="FFFFFF", bold=True) 
                alineacion_centro = Alignment(horizontal="center", vertical="center")
                borde_fino = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                for col_num, nombre_columna in enumerate(df.columns):
                    celda = worksheet.cell(row=1, column=col_num + 1)
                    celda.fill = color_enc
                    celda.font = fuente_enc
                    celda.alignment = alineacion_centro
                    celda.border = borde_fino
                    worksheet.column_dimensions[get_column_letter(col_num + 1)].width = max(len(str(nombre_columna)), df[nombre_columna].astype(str).map(len).max()) + 2

                for fila in worksheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
                    for celda in fila:
                        celda.alignment = alineacion_centro
                        celda.border = borde_fino
                        
            st.download_button(label="📥 Descargar Reporte (.xlsx)", data=buffer.getvalue(), file_name="reporte_finca_v3.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
            st.markdown("---")
            if st.button("🚪 Cerrar Sesión"):
                st.session_state['autorizado'] = False
                st.rerun()

    else:
        st.info("La base de datos está vacía.")
except Exception as error:
    st.error(f"Error de conexión: {error} 🔴")